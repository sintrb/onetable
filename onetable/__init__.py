# -*- coding: UTF-8 -*
'''
Created on 2019/06/22

@author: Robin
'''

__version__ = '0.0.4'


class UniTable(object):
    defaultValue = ''
    defaultStyle = ''

    def __init__(self):
        from collections import OrderedDict
        self.styles = OrderedDict()
        self.widths = OrderedDict()
        self.lastRow = -1
        self.lastCol = -1
        self.maxRow = 0
        self.maxCol = 0
        self.data = {}

    def getValue(self, cell):
        value = cell.get('value', self.defaultValue)
        if callable(value):
            value = value(cell)
        return value

    def getStyle(self, cell):
        style = cell.get('style', self.defaultStyle)
        if callable(style):
            style = style(cell)
        if type(style) == dict:
            print(style, cell)
        return style

    def setWidth(self, col, width):
        self.widths[col] = width

    def setWidths(self, *widths):
        for c, w in enumerate(widths):
            self.setWidth(c, w)

    def getCellWidth(self, row, col, colspan):
        w = self.widths.get(col, 0)
        for c in range((colspan or 1) - 1):
            w += self.widths.get(col + c + 1, 0)
        return w

    def setStyle(self, name, style):
        self.styles[name] = style

    def setStyles(self, *styletuple):
        for name, style in styletuple:
            self.setStyle(name, style)

    def writeCellWithSpan(self, row, col, rowspan, colspan, value, style=None):
        if rowspan == None:
            raise Exception()
        for r in range(rowspan):
            for c in range(colspan):
                ix = (row + r), (col + c)
                if ix in self.data:
                    raise Exception(u'cell {ix} conflict with value: {val}'.format(ix=ix, val=value))
                cell = {
                    'value': value,
                    'style': style,
                    'span': r or c
                }
                if not cell['span']:
                    cell['style'] = style
                    cell['rowspan'] = rowspan
                    cell['colspan'] = colspan
                self.data[ix] = cell
        self.lastRow = row + rowspan - 1
        self.lastCol = col + colspan - 1
        self.maxRow = max(self.maxRow, self.lastRow)
        self.maxCol = max(self.maxCol, self.lastCol)

    def writeCellWithColSpan(self, row, col, span, value, style=None):
        return self.writeCellWithSpan(row, col, rowspan=1, colspan=span, value=value, style=style)

    def writeCell(self, row, col, value, style=None):
        return self.writeCellWithSpan(row, col, rowspan=1, colspan=1, value=value, style=style)

    def writeNextCol(self, value, colspan=1, style=None):
        return self.writeCellWithColSpan(max(self.lastRow, 0), max(self.lastCol + 1, 0), colspan, value=value,
                                         style=style)

    def writeNextRow(self, value, colspan=1, style=None):
        return self.writeCellWithColSpan(max(self.lastRow + 1, 0), 0, colspan, value=value, style=style)

    def writeNextRowCols(self, values, colspan=1, style=None):
        self.lastRow += 1
        self.lastCol = -1
        return self.writeNextCols(values, colspan, style)

    def writeNextCols(self, values, colspan=1, style=None):
        cells = []
        for v in values:
            if type(v) == dict:
                value = v if type(v) != dict else v.get('value')
                cs = v.get('colspan', colspan)
                st = v.get('style') or style
            elif type(v) in [list, tuple]:
                value = v[0] if len(v) > 0 else None
                cs = colspan  # v[1] if len(v) > 1 else colspan
                st = style  # v[2] if len(v) > 2 else style
                for sv in v[1:]:
                    if isinstance(sv, str):
                        st = sv
                    elif type(sv) == int:
                        cs = sv
            else:
                value = v
                cs = colspan
                st = style
            cells.append(self.writeNextCol(value, cs, st))
        return cells

    def getCellByRowCol(self, row, col):
        return self.data.get((row, col), )

    _status = []

    def pushStatus(self):
        self._status.append((self.lastRow, self.lastCol), )

    def popStatus(self):
        self.lastRow, self.lastCol = self._status.pop()


class DefualtTable(UniTable):
    def __init__(self, style=None):
        self.defaultStyleDict = {'fontSize': 13, 'font': u'SimSun', 'background': 'white'}
        if style:
            self.defaultStyleDict.update(style)
        UniTable.__init__(self)
        self.setStyle('default', self.defaultStyleDict)

    def getStyle(self, cell):
        style = UniTable.getStyle(self, cell) or 'default'
        return style

    def setStyle(self, name, style):
        for k, v in self.defaultStyleDict.items():
            if k not in style:
                style[k] = v
        return UniTable.setStyle(self, name, style)


def write_complexhead(table, archcolumns, style='headstyle', rowindex=0):
    columns = []

    def wrapColsSpan(cols, rowspan=0):
        rowspan += 1
        if not cols:
            return 1, rowspan
        colspan = 0
        for c in cols:
            c['colspan'], _ = wrapColsSpan(c.get('children'), rowspan)
            colspan += c['colspan']
            if not c.get('children'):
                columns.append(c)
        return colspan, rowspan

    rows = []

    def wrapColsRow(cols, row=0):
        if not cols:
            return
        if row >= len(rows):
            rows.append([])
        for c in cols:
            rows[row].append(c)
            wrapColsRow(c.get('children'), row + 1)

    def wrapColsCol(cols, col=0):
        if not cols:
            return
        for c in cols:
            c['col'] = col
            wrapColsCol(c.get('children'), col)
            col += c.get('colspan')

    wrapColsSpan(archcolumns)
    wrapColsCol(archcolumns)
    wrapColsRow(archcolumns)

    def wrapCell(c):
        nc = c.copy()
        nc['value'] = c['title']
        return nc

    rowindex = 0
    for cols in rows:
        coloffset = 0
        for c in cols:
            table.writeCellWithSpan(rowindex,
                                    coloffset + c.get('col', 0),
                                    c.get('rowspan', 1),
                                    c.get('colspan', 1),
                                    c['title'],
                                    style=c.get('style', style))
        rowindex += 1
    return columns


def gen_dataitemstable(items, columns, sumfields=None, styles=[], cellwidth=0, table=None, complexhead=False,
                       rowstyle=None, wrapsum=None):
    '''
    @items: iterateable object, ex: list, set, tuple...
    @colums: the list of colum define, ex: [{'title':'the title', 'field':'field or key of item', 'getter':'the getter of value, def(item)', 'transfer':'the transfer of value, def(item)', 'style':'style name or getter if style name'}, ...]
    @sumfields: the list of field which will calculate sum
    @styles: the list of style define, ex: [('big', {'fontSize':100})]
    @cellwidth: the default width for each column
    @table: UniTable object
    @complexhead: write with complex head title or not
    @rowstyle: the style name or getter function for rows, getter function sign with def(index, item), the head row is (-1, None), the sum row is (-2, None)
    '''
    tb = table or DefualtTable()
    tb.setStyles(
        ('nametitle', {'align': 'center', 'fontSize': 13, 'bold': True}),
    )
    tb.setStyles(*styles)
    tb.setWidths(*[col.get('width', cellwidth) for col in columns])
    if complexhead:
        columns = write_complexhead(tb, columns, style=rowstyle and (
            callable(rowstyle) and rowstyle(-1, None) or not callable(rowstyle)) or 'nametitle')
    else:
        tb.writeNextRowCols([col.get('title') for col in columns], style=rowstyle and (
            callable(rowstyle) and rowstyle(-1, None) or not callable(rowstyle)) or 'nametitle')
    sumdict = {
        f: 0
        for f in sumfields or []
    }

    def safefloat(v):
        try:
            return float(v)
        except:
            return 0

    def getval(item, col):
        field = col["field"]
        val = None
        if type(item) == dict:
            if field in item:
                val = item[field]
        elif hasattr(item, field):
            val = getattr(item, field)
        if col.get('getter'):
            val = col['getter'](item)
        if col.get('transfer'):
            val = col.get('transfer')(val)
        st = col.get('style')
        if st and callable(st):
            st = st(item, col)
        if field in sumdict and val:
            sumdict[field] += safefloat(val or 0)
        return {
            'value': val,
            'style': st
        }

    for index, item in enumerate(items):
        style = rowstyle
        if callable(style):
            style = style(index, item)
        tb.writeNextRowCols([getval(item, col) for col in columns], style=style or 'default')
    if sumfields:
        if wrapsum:
            sumdict = wrapsum(sumdict) or sumfields
        tb.writeNextRowCols([{'value': sumdict.get(col['field'], '-'), 'style': col.get('style')} for col in columns],
                            style=rowstyle and (callable(rowstyle) and rowstyle(-2, None) or not callable(
                                rowstyle)) or 'default')

    return tb


class TableRender(object):
    fontSizeRatio = 1
    widthRatio = 1

    def __init__(self, scale=1):
        self.fontSizeRatio = self.fontSizeRatio * scale
        self.widthRatio = self.widthRatio * scale

    def render(self, tab):
        raise Exception(u'Unimplement !')

    def save(self, tab, out):
        '''out is file like object, out.write must support'''
        raise Exception(u'Unimplement !')


class HtmlRender(TableRender):
    '''
    Using to render to html table
    '''

    fontSizeRatio = 1.1
    widthRatio = 1.3

    def renderTableTag(self, tab):
        from decimal import Decimal
        trs = []
        for r in range(tab.maxRow + 1):
            tds = []
            for c in range(tab.maxCol + 1):
                cell = tab.getCellByRowCol(r, c) or {}
                if cell and cell.get('span'):
                    continue
                value = tab.getValue(cell)  # cell.get('value')
                #                 print value,type(value)
                if type(value) in [float, Decimal]:
                    #                     print value
                    if int(value) == value:
                        value = int(value)
                rowspan = cell.get('rowspan', 1)
                colspan = cell.get('colspan', 1)
                style = tab.getStyle(cell)
                width = tab.getCellWidth(r, c, cell.get('colspan'))
                attr = {
                    'rowspan': rowspan if rowspan > 1 else '',
                    'colspan': colspan if colspan > 1 else '',
                    'class': style,
                    'width': int(width * self.widthRatio),
                }
                if value and isinstance(value, str):
                    value = value.replace('\n', '<br/>')
                td = u'<td {attrs}>{value}</td>'.format(
                    value=value,
                    attrs=u' '.join([u'%s="%s"' % (k, v) for k, v in attr.items() if v])
                )
                tds.append(td)
            tr = u'<tr>\n\t\t{tds}\n\t</tr>'.format(tds=u'\n\t\t'.join(tds))
            trs.append(tr)
        tabhtml = u'<table>\n\t{trs}\n</table>'.format(trs=u'\n\t'.join(trs))
        return tabhtml

    def renderStyleTag(self, tab, prex='td.'):
        from collections import OrderedDict
        cssClasses = []
        colormap = {
            'red': 'red',
            'black': 'black',
            'blue': 'blue',
            'green': 'green',
            'white': 'white',
            'lavender': '#C572DA',
            'lime': '#63C544',
            'light_orange': '#F49638',
            'coral': 'coral',
            'yellow': 'yellow',
            'purple': '#ba68c8',
            'light_green': 'lightgreen',
            'pink': '#ff60a2',
            'lightred': '#da1019',
            'lightpink': '#FFB6C1',
            'ice_blue': '#f4e0fb',
            'pale_blue': '#bde0fb',
        }
        for name, style in tab.styles.items():
            cssd = OrderedDict()
            if 'color' in style:
                cssd['color'] = colormap.get(style['color'], style['color'] or 'black')
            if 'background' in style:
                cssd['background'] = colormap.get(style['background'], style['background'] or 'white')
            if 'align' in style:
                cssd['text-align'] = style['align']
            if 'fontSize' in style:
                cssd['font-size'] = u'{0}px'.format((style['fontSize'] * self.fontSizeRatio))
            if 'height' in style:
                cssd['height'] = u'{0}px'.format((style['height'] * self.widthRatio))
            if style.get('bold'):
                cssd['font-weight'] = 'bold'
            if style.get('font'):
                cssd['font-family'] = style['font']
            css = u'{' + u''.join(
                [u'{key}:{value};'.format(key=key, value=value) for key, value in cssd.items()]) + u'}'
            cssClasses.append(u'\t{prex}{name}{css}'.format(prex=prex, name=name, css=css))
        return u'<style type="text/css">\n{csses}\n</style>'.format(csses=u'\n'.join(cssClasses))

    def render(self, tab):
        return self.renderStyleTag(tab), self.renderTableTag(tab)

    def save(self, tab, out):
        st, tb = self.render(tab)
        html = u'''<html><head>{st}</head><body>{tb}</body></html>'''.format(st=st, tb=tb)
        out.write(html.encode('utf-8'))


class XlsRender(TableRender):
    '''
    Using xlwt to render to xls
    '''

    fontSizeRatio = 15
    widthRatio = 28

    def render(self, tab, book=None, sheet=None):
        from xlwt import Workbook, Worksheet, easyxf, Style
        book = Workbook() if not book else book
        transheetname = lambda v: v.replace(':', '-').replace('/', '-').replace('\\', '-').replace('?', '-').replace(
            '*', '-')[0:30]
        ws = sheet if isinstance(sheet, Worksheet) else book.add_sheet(transheetname(sheet or u'Sheet1'))
        stylemap = {}
        colormap = {
            'purple': 'lavender',
            'red': 'pink',
            'lightred': 'red',
            'gray': 'gray40',
            'lightpink': 'rose',
        }
        borders = u'borders: left thin, top thin, right thin, bottom thin;'
        for name, style in tab.styles.items():
            xlsst = u'' if 'border' in style and style['border'] == None else borders
            # font
            color = style.get('color', 'black')
            xlsst += u'font:color {color}, name {font}, height {height}, bold {bold};'.format(
                font=style.get('font', u'宋体'),
                color=colormap.get(color, color),
                height=int(style.get('fontSize', 12) * self.fontSizeRatio),
                bold=style.get('bold', False) and 'on' or 'off'
            )
            # align
            xlsst += u'align:horiz {align}, vertical center, wrap {wrap};'.format(align=style.get('align', 'left'),
                                                                                  wrap=style.get('wrap',
                                                                                                 False) and 'on' or 'off')

            # background
            background = style.get('background', 'white')
            xlsst += u'pattern:pattern solid, fore_colour {background};'.format(
                background=colormap.get(background, background))
            xst = easyxf(xlsst)
            #             xst.num_format_str = "#0.00"
            stylemap[name] = xst

        for r in range(tab.maxRow + 1):
            height_mismatch = False
            for c in range(tab.maxCol + 1):
                cell = tab.getCellByRowCol(r, c) or {}
                if cell and cell.get('span'):
                    continue
                value = tab.getValue(cell)  # cell.get('value')
                rowspan = cell.get('rowspan', 1)
                colspan = cell.get('colspan', 1)
                style = tab.getStyle(cell)
                if style and style not in stylemap:
                    raise Exception(u'%s not in styles' % style)
                if isinstance(value, str):
                    value = u'%s' % value
                if rowspan > 1 or colspan > 1:
                    ws.write_merge(r, r + rowspan - 1, c, c + colspan - 1, value,
                                   style and stylemap.get(style) or Style.default_style)
                else:
                    ws.write(r, c, value, style and stylemap.get(style) or Style.default_style)
                if c == 0:
                    styled = tab.styles.get(style)
                    if styled and 'height' in styled:
                        height_mismatch = True
                        ws.row(r).height = int(styled['height'] * self.widthRatio * 0.9)
            ws.row(r).height_mismatch = height_mismatch
        for col, width in tab.widths.items():
            if not width:
                continue
            ws.col(col).width = int(width * self.widthRatio)
        return book, ws

    def save(self, tab, out):
        book, _ = self.render(tab)
        book.save(out)


class XlsxRender(TableRender):
    '''
    Using openpyxl to render to xlsx, see the document http://openpyxl.readthedocs.io/en/default/
    '''
    fontSizeRatio = 0.8
    widthRatio = 0.108

    def render(self, tab):
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = u'Sheet1'

        defaultside = Side(border_style="thin", color="000000")
        defaultborder = Border(defaultside, defaultside, defaultside, defaultside)

        stylemap = {}
        colormap = {
            'red': 'FFFF0000',
            'black': 'FF000000',
            'blue': 'FF0000FF',
            'green': 'FF00FF00',
            'white': 'FFFFFFFF',
            'lavender': 'FFC572DA',
            'lime': 'FF63C544',
            'light_orange': 'FFF49638',
            'coral': 'FFFB4D50',
        }

        def transcolor(c):
            return colormap.get(c)

        for name, style in tab.styles.items():
            border = None if 'border' in style and style['border'] == None else defaultborder
            font = Font(name=style.get('font'),
                        size=int(style.get('fontSize', 12) * self.fontSizeRatio),
                        bold=style.get('bold', False),
                        italic=False,
                        vertAlign=None,
                        underline='none',
                        strike=False,
                        color=transcolor(style.get('color', 'FF000000')))
            alignment = Alignment(horizontal=style.get('align', 'left'),
                                  vertical='center',
                                  text_rotation=0,
                                  wrap_text=True,
                                  shrink_to_fit=True,
                                  indent=0)
            styled = {
                'border': border,
                'font': font,
                'alignment': alignment
            }
            stylemap[name] = styled

        def applystyled(cs, styled):
            cs.border = styled.get('border')
            cs.font = styled.get('font')
            cs.alignment = styled.get('alignment')

        for r in range(tab.maxRow + 1):
            for c in range(tab.maxCol + 1):
                cell = tab.getCellByRowCol(r, c) or {}
                if cell and cell.get('span'):
                    continue
                value = tab.getValue(cell)  # cell.get('value')
                rowspan = cell.get('rowspan', 1)
                colspan = cell.get('colspan', 1)
                style = tab.getStyle(cell)
                if style and style not in stylemap:
                    raise Exception(u'%s not in styles' % style)
                styled = stylemap.get(style) or {}
                if isinstance(value, str):
                    value = u'%s' % value
                col = c + 1
                row = r + 1
                cs = ws.cell(column=col, row=row, value=value)
                if rowspan > 1 or colspan > 1:
                    ws.merge_cells(start_row=row, start_column=col, end_row=row + rowspan - 1,
                                   end_column=col + colspan - 1)
                    for ri in range(rowspan):
                        for ci in range(colspan):
                            cs = ws.cell(row=row + ri, column=col + ci)
                            applystyled(cs, styled)
                else:
                    applystyled(cs, styled)
        for col, width in tab.widths.items():
            if not width:
                continue
            ws.column_dimensions[get_column_letter(col + 1)].width = int(width * self.widthRatio)

        ws.print_options.horizontalCentered = True
        ws.print_options.verticalCentered = False
        ws.page_margins.left = ws.page_margins.right = 0
        #         ws.page_margins.top = ws.page_margins.bottom = 0
        return wb, ws

    def save(self, tab, out):
        book, _ = self.render(tab)
        book.save(out)


class PdfRender(TableRender):
    '''
    render to pdf
    '''
    fontSizeRatio = 0.5
    widthRatio = 0.6

    def render(self, tab):
        from reportlab.pdfgen.canvas import Canvas
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.cidfonts import UnicodeCIDFont
        pdfmetrics.registerFont(UnicodeCIDFont('STSong-Light'))
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle

        stylemap = {}
        for name, style in tab.styles.items():
            stylemap[name] = style
        colormap = {
            'red': colors.red,
            'green': colors.green,
            'blue': colors.blue,
            'black': colors.black,
            'white': colors.white,
        }

        def transcolor(c):
            return colormap.get(c) or colors.black

        tabdata = []
        styles = [
            ('FONTNAME', (0, 0), (-1, -1), 'STSong-Light'),
            #             ('FONTNAME', (1, 1), (-1, -1), 'STSong-Light'),
            ('GRID', (0, 0), (-1, -1), 0.5, '#FF000000'),
            ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
        ]
        for r in range(tab.maxRow + 1):
            cols = []
            for c in range(tab.maxCol + 1):
                cell = tab.getCellByRowCol(r, c) or {}
                if cell and cell.get('span'):
                    cols.append('')
                    continue
                value = tab.getValue(cell)  # cell.get('value')
                rowspan = cell.get('rowspan', 1)
                colspan = cell.get('colspan', 1)
                style = tab.getStyle(cell)
                if style and style not in stylemap:
                    raise Exception(u'%s not in styles' % style)
                styled = stylemap.get(style) or {}
                if isinstance(value, str):
                    value = u'%s' % value
                col = c  # + 1
                row = r  # + 1
                cix = (r, c)

                styles += [
                    ('FONTSIZE', cix, (0, 0), int(styled.get('fontSize', 12) * self.fontSizeRatio)),
                    ('ALIGN', cix, (0, 0), styled.get('align', 'left').upper()),
                    ('TEXTCOLOR', cix, (0, 0), transcolor(styled.get('color', 'black')))
                ]
                if rowspan > 1 or colspan > 1:
                    styles.append(
                        ('SPAN', (col, row), (col + colspan - 1, row + rowspan - 1))
                    )
                cols.append(value)
            tabdata.append(cols)
        widths = [20 for _ in range(len(tab.widths))]
        for col, width in tab.widths.items():
            if len(widths) > col >= 0 and width > 0:
                widths[col] = int(width * self.widthRatio)
        pdftab = Table(tabdata, colWidths=(widths or None))
        pdftab.setStyle(TableStyle(styles))
        return pdftab

    def save(self, tab, out):
        from reportlab.platypus import SimpleDocTemplate
        from reportlab.lib.pagesizes import A4
        story = self.render(tab)
        doc = SimpleDocTemplate(out, showBoundary=0, leftMargin=0, rightMargin=0, topMargin=0, bottomMargin=0, pagesize=A4)
        doc.build([story])


class DataGridRender(HtmlRender):
    '''
    render to easy datagrid
    '''

    def render(self, tab):
        columns = []
        rows = []
        for r in range(tab.maxRow + 1):
            tds = []
            data = {}
            for c in range(tab.maxCol + 1):
                field = u'c%d' % c
                cell = tab.getCellByRowCol(r, c) or {}
                rowspan = cell.get('rowspan', 1)
                colspan = cell.get('colspan', 1)
                style = tab.getStyle(cell)
                if r == 0:
                    # title
                    width = tab.getCellWidth(r, c, cell.get('colspan'))
                    columns.append({
                        'field': field,
                        'title': cell['value'],
                        'width': width,
                    })
                elif 'value' in cell:
                    if not style:
                        data[field] = cell['value']
                    else:
                        data[field] = {
                            'v': cell['value'],
                            's': style
                        }
            if data:
                rows.append(data)
        return self.renderStyleTag(tab, prex='.grid-style-'), {
            'data': {
                'rows': rows
            },
            'columns': [columns, ],
        }


class CSVRender(TableRender):
    def render(self, tab):
        lines = []
        for r in range(tab.maxRow + 1):
            ls = []
            for c in range(tab.maxCol + 1):
                cell = tab.getCellByRowCol(r, c) or {}
                if cell and cell.get('span'):
                    ls.append('')
                    continue
                value = tab.getValue(cell)
                value = u'%s' % value
                if '\n' in value:
                    value = value.replace('"', '""')
                    value = u'"%s"' % value
                ls.append(value)
            lines.append(u','.join(ls))
        return u'\n'.join(lines)

    def save(self, tab, out):
        content = self.render(tab)
        out.write(content.encode('utf-8'))


OneTable = UniTable  # the old name is unitable, but pypi was registered.
